#!/usr/bin/env python3
"""
Rutelister RTF → Excel Konverter - Web App
==========================================
Kør med: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import re
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import zipfile
from datetime import date

st.set_page_config(
    page_title="Rutelister Konverter",
    page_icon="🚚",
    layout="wide"
)

# Custom CSS for moderne look
st.markdown("""
<style>
    /* Mørkt tema styling */
    .main {
        background-color: #1e1e2e;
    }

    /* Header styling */
    .main-header {
        background: linear-gradient(135deg, #7c3aed 0%, #4f46e5 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        color: white;
    }

    .main-header h1 {
        margin: 0;
        font-size: 2.5rem;
    }

    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
    }

    /* Upload zone styling */
    .upload-zone {
        border: 2px dashed #7c3aed;
        border-radius: 12px;
        padding: 3rem;
        text-align: center;
        background: #2d2d44;
        margin: 1rem 0;
    }

    .upload-zone:hover {
        border-color: #8b5cf6;
        background: #363652;
    }

    /* Fil info cards */
    .file-card {
        background: #2d2d44;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        border-left: 4px solid #7c3aed;
    }

    /* Success box */
    .success-box {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
    }

    /* Stats cards */
    .stat-card {
        background: #2d2d44;
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
    }

    .stat-number {
        font-size: 2.5rem;
        font-weight: bold;
        color: #7c3aed;
    }

    .stat-label {
        color: #94a3b8;
        font-size: 0.9rem;
    }

    /* Farve legend */
    .color-legend {
        display: flex;
        flex-wrap: wrap;
        gap: 1rem;
        margin: 1rem 0;
    }

    .color-item {
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }

    .color-box {
        width: 20px;
        height: 20px;
        border-radius: 4px;
    }

    /* Fjern fade-effekt ved rerun */
    [data-testid="stAppViewContainer"],
    [data-testid="stVerticalBlock"],
    .stApp, .main, .block-container,
    [data-testid="stMainBlockContainer"] {
        opacity: 1 !important;
        transition: none !important;
    }
    div[data-stale="true"] {
        opacity: 1 !important;
    }

    /* Streamlit file uploader styling */
    [data-testid="stFileUploader"] {
        background: #2d2d44;
        border-radius: 12px;
        padding: 1rem;
    }

    [data-testid="stFileUploader"] > div {
        border: 2px dashed #7c3aed !important;
        border-radius: 12px !important;
    }
</style>
""", unsafe_allow_html=True)


# ============= RTF PARSING FUNKTIONER =============

HEADER_FOOTER_PAT = re.compile(
    r'(Hasselager FVT|TUR START|TRIP|PAUSE|LÆSSEPORT|HOSTRUTE|VOGNNUMMER|ÅBNE - LUKKE|'
    r'STARTTID|HJEMKOMSTTID|SLUTTID|ROUTEDATE|Udskrevet:|Side\s+\d+\s+af\s+\d+)', re.IGNORECASE
)


def rtf_to_text(rtf: str) -> str:
    """Konverter RTF til plain text"""
    def uni_sub(m):
        codepoint = int(m.group(1))
        if codepoint < 0:
            codepoint = 65536 + codepoint
        try:
            ch = chr(codepoint)
        except ValueError:
            ch = ' '
        return ch
    text = re.sub(r'\\u(-?\d+)\??', uni_sub, rtf)

    def hex_sub(m):
        try:
            return bytes.fromhex(m.group(1)).decode('latin-1', errors='ignore')
        except Exception:
            return ''
    text = re.sub(r"\\'([0-9a-fA-F]{2})", hex_sub, text)

    text = re.sub(r'\\par[d]?', '\n', text)
    text = re.sub(r'\\line', '\n', text)
    text = re.sub(r'\\cell', '\n', text)
    text = re.sub(r'\\row', '\n', text)
    text = re.sub(r'\\tab', '    ', text)
    text = re.sub(r'\\[a-zA-Z]+\d* ?', ' ', text)
    text = text.replace('{', ' ').replace('}', ' ')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n[ \t]+', '\n', text)
    text = re.sub(r'[ \t]+\n', '\n', text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text


def split_pages(lines):
    """Split linjer i sider baseret på sidefod"""
    pages = []
    current = []
    for ln in lines:
        current.append(ln)
        if re.search(r"Side\s+\d+\s+af\s+\d+", ln):
            pages.append(current)
            current = []
    if current:
        pages.append(current)
    return pages


def find_meta(page_lines):
    """Find metadata fra side"""
    host_re = re.compile(r'HOSTRUTE:\s*(\d+)')
    port_re = re.compile(r'LÆSSEPORT:\s*(\d+)')
    start_re = re.compile(r'STARTTID:\s*([0-2]?\d:\d{2})')
    slut_re = re.compile(r'SLUTTID:\s*([0-2]?\d:\d{2})')
    afr_re = re.compile(r'AFREGNINGSTID:\s*(\d+)')

    rutenummer = portnummer = starttid = sluttid = afregningstid = None
    for ln in page_lines:
        m = host_re.search(ln)
        rutenummer = m.group(1) if m else rutenummer
        m = port_re.search(ln)
        portnummer = m.group(1) if m else portnummer
        m = start_re.search(ln)
        starttid = m.group(1) if m else starttid
        m = slut_re.search(ln)
        sluttid = m.group(1) if m else sluttid
        m = afr_re.search(ln)
        afregningstid = m.group(1) if m else afregningstid
    return rutenummer, portnummer, starttid, sluttid, afregningstid


def find_street_and_post(page_lines, start_idx, lookahead_depth=12):
    """Find adresse og postnummer"""
    street = None
    postnr = None
    by = None

    n = len(page_lines)
    lookahead = []
    for k in range(1, lookahead_depth + 1):
        if start_idx + k < n:
            lookahead.append(page_lines[start_idx + k])

    for s in lookahead:
        if not s or HEADER_FOOTER_PAT.search(s):
            continue
        if re.search(r'\b\d{4}\b', s):
            continue
        street = s.strip()
        break

    for s in lookahead:
        pm = re.search(r'(\d{4})\s+(.+)$', s)
        if pm:
            postnr = pm.group(1)
            by = pm.group(2).strip()
            break

    if street is None:
        for s in lookahead:
            pm2 = re.search(r'(.+?)\s+(\d{4})\s+(.+)$', s)
            if pm2 and not HEADER_FOOTER_PAT.search(s):
                street = pm2.group(1).strip()
                postnr = pm2.group(2)
                by = pm2.group(3).strip()
                break

    return street, postnr, by


def parse_page(page_lines):
    """Parse en enkelt side og udtræk rækker"""
    stop_line_re = re.compile(
        r'^(?P<id>\d{5})\s+(?P<name>.+?)\s+'
        r'(?:\d{1,2}:\d{2})\s*-\s*(?:\d{1,2}:\d{2})'
        r'(?:\s+\d+){1,8}\s+'
        r'(?P<ank>\d{1,2}:\d{2})\s+(?P<afg>\d{1,2}:\d{2})\b'
    )

    rutenummer, portnummer, starttid, sluttid, afregningstid = find_meta(page_lines)

    rows = []
    for i, ln in enumerate(page_lines):
        sm = stop_line_re.match(ln)
        if not sm:
            continue

        name_raw = sm.group('name').strip()
        ank = sm.group('ank')

        street, postnr, by = find_street_and_post(page_lines, i, lookahead_depth=12)

        if any(x and 'Hasselager' in x for x in [name_raw, street, by]):
            continue

        butik = re.sub(r'\s+[A]\s*\d*$', '', name_raw).strip()

        afr_hours = None
        if afregningstid:
            try:
                afr_hours = round(float(afregningstid) / 60.0, 2)
            except:
                afr_hours = None

        rows.append({
            "Butiksnavn": butik,
            "Adresse": street,
            "Postnr": postnr,
            "By": by,
            "Ankomst": ank,
            "Rutenummer": rutenummer,
            "Portnummer": portnummer,
            "Starttid": starttid,
            "Sluttid": sluttid,
            "Afregningstid (timer)": afr_hours
        })
    return rows


def process_rtf_file(file_content: bytes, filename: str) -> tuple:
    """Behandl en RTF fil og returner DataFrame + statistik"""
    raw = file_content.decode("latin-1", errors="ignore")
    text = rtf_to_text(raw)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    pages = split_pages(lines)

    all_rows = []
    for page in pages:
        all_rows.extend(parse_page(page))

    df = pd.DataFrame(all_rows)
    return df, len(all_rows)


def create_excel(df: pd.DataFrame, koerselsdato: str = None) -> BytesIO:
    """Opret farvekodede Excel fil med fed skrift på første række af hvert rutenummer"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ruter"

    # Tilføj kørselsdato kolonne hvis angivet
    if koerselsdato and 'Kørselsdato' not in df.columns:
        df = df.copy()
        df['Kørselsdato'] = koerselsdato

    # Rutenummer først, derefter resten - nu med Kørselsdato
    cols = ["Kørselsdato", "Rutenummer", "Portnummer", "Butiksnavn", "Adresse", "Postnr", "By",
            "Ankomst", "Starttid", "Sluttid", "Afregningstid (timer)"]
    ws.append(cols)

    # Header style
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Append data og find første/sidste række for hvert rutenummer
    route_rows = {}  # rutenummer -> [første_row, sidste_row]
    prev_rutenummer = None

    if not df.empty:
        for idx, row in df[cols].iterrows():
            ws.append(row.tolist())
            current_row = ws.max_row
            current_rutenummer = row["Rutenummer"]

            # Track første og sidste række per rute
            if current_rutenummer != prev_rutenummer:
                # Afslut forrige rute
                if prev_rutenummer is not None and prev_rutenummer in route_rows:
                    route_rows[prev_rutenummer][1] = current_row - 1
                # Start ny rute
                route_rows[current_rutenummer] = [current_row, current_row]
                prev_rutenummer = current_rutenummer
            else:
                # Opdater sidste række for nuværende rute
                if current_rutenummer in route_rows:
                    route_rows[current_rutenummer][1] = current_row

    # Farver
    YELLOW = "FFF9C4"
    BLUE = "BBDEFB"
    GREEN = "C8E6C9"
    PURPLE = "E1BEE7"
    GREY = "ECEFF1"
    ORANGE = "FFE0B2"  # Kørselsdato farve

    fill_map = {
        "Kørselsdato": ORANGE,
        "Butiksnavn": YELLOW, "Adresse": YELLOW, "Postnr": YELLOW, "By": YELLOW,
        "Ankomst": BLUE,
        "Rutenummer": GREEN,
        "Portnummer": PURPLE,
        "Starttid": GREY, "Sluttid": GREY, "Afregningstid (timer)": GREY
    }
    col_idx = {name: idx + 1 for idx, name in enumerate(cols)}
    max_row = ws.max_row

    for name, color in fill_map.items():
        idx = col_idx[name]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=idx).fill = fill

    # Definer kant-styles
    thick_border = Side(style='medium', color='000000')
    thin_border = Side(style='thin', color='AAAAAA')

    # Anvend fed skrift på første række + kanter omkring hver rute-gruppe
    for rutenummer, (first_row, last_row) in route_rows.items():
        # Fed skrift på første række
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=first_row, column=c)
            cell.font = Font(bold=True)

        # Kant omkring hele rute-gruppen
        for r in range(first_row, last_row + 1):
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)

                # Bestem kanterne for denne celle
                top = thick_border if r == first_row else thin_border
                bottom = thick_border if r == last_row else thin_border
                left = thick_border if c == 1 else thin_border
                right = thick_border if c == len(cols) else thin_border

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Freeze header og autosize
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[column].width = min(60, max(12, max_length + 2))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_zip_with_all_files(files_data: list) -> BytesIO:
    """Opret ZIP fil med alle Excel filer"""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename, excel_data in files_data:
            zf.writestr(filename, excel_data.getvalue())
    zip_buffer.seek(0)
    return zip_buffer


# ============= STREAMLIT UI =============

def main():
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>🚚 Rutelister Konverter</h1>
        <p>Konverter RTF-rutelister til farvekodede Excel-ark</p>
    </div>
    """, unsafe_allow_html=True)

    # Sidebar med info
    with st.sidebar:
        st.header("📋 Farve-guide")
        st.markdown("""
        | Farve | Kolonner |
        |-------|----------|
        | 🟧 Orange | Kørselsdato |
        | 🟨 Gul | Butik, Adresse, Postnr, By |
        | 🔵 Blå | Ankomst |
        | 🟢 Grøn | Rutenummer |
        | 🟣 Lilla | Portnummer |
        | ⬜ Grå | Starttid, Sluttid, Afregning |
        """)

        st.divider()

        st.header("ℹ️ Sådan bruger du")
        st.markdown("""
        1. **Upload** en eller flere RTF-filer
        2. **Se preview** af data
        3. **Download** Excel-filer
        """)

    # Hovedindhold
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("📁 Upload RTF-filer")
        uploaded_files = st.file_uploader(
            "Træk filer hertil eller klik for at vælge",
            type=['rtf'],
            accept_multiple_files=True,
            help="Du kan uploade flere RTF-filer på én gang"
        )

    with col2:
        st.subheader("📅 Kørselsdato")
        koerselsdato = st.date_input(
            "Vælg dato for kørslen",
            value=date.today(),
            format="DD-MM-YYYY",
            help="Denne dato tilføjes til alle rækker i Excel-filen"
        )
        koerselsdato_str = koerselsdato.strftime("%d-%m-%Y")

        st.subheader("📊 Status")
        if uploaded_files:
            st.metric("Filer uploadet", len(uploaded_files))
        else:
            st.info("Ingen filer uploadet endnu")

    # Behandl filer
    if uploaded_files:
        st.divider()
        st.subheader("🔄 Behandling")

        progress_bar = st.progress(0)
        status_text = st.empty()

        all_results = []
        total_rows = 0

        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Behandler: {uploaded_file.name}")
            progress_bar.progress((i + 1) / len(uploaded_files))

            try:
                content = uploaded_file.read()
                df, row_count = process_rtf_file(content, uploaded_file.name)

                excel_data = create_excel(df, koerselsdato_str)
                output_name = Path(uploaded_file.name).stem + "_farvestruktur.xlsx"

                all_results.append({
                    'filename': uploaded_file.name,
                    'output_name': output_name,
                    'df': df,
                    'row_count': row_count,
                    'excel_data': excel_data
                })
                total_rows += row_count

            except Exception as e:
                st.error(f"Fejl ved behandling af {uploaded_file.name}: {str(e)}")

        status_text.text("Færdig!")
        progress_bar.progress(1.0)

        # Resultater
        if all_results:
            st.divider()

            # Statistik
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📄 Filer behandlet", len(all_results))
            with col2:
                st.metric("📝 Total rækker", total_rows)
            with col3:
                avg_rows = total_rows // len(all_results) if all_results else 0
                st.metric("📈 Gns. rækker/fil", avg_rows)

            st.divider()

            # Download sektion
            st.subheader("⬇️ Download")

            # Download alle som ZIP
            if len(all_results) > 1:
                files_for_zip = [(r['output_name'], r['excel_data']) for r in all_results]
                zip_data = create_zip_with_all_files(files_for_zip)
                st.download_button(
                    label="📦 Download alle som ZIP",
                    data=zip_data,
                    file_name="rutelister_excel.zip",
                    mime="application/zip",
                    use_container_width=True,
                    type="primary"
                )
                st.markdown("---")

            # Individuelle downloads
            for result in all_results:
                with st.expander(f"📄 {result['filename']} → {result['output_name']} ({result['row_count']} rækker)", expanded=False):
                    col1, col2 = st.columns([3, 1])

                    with col1:
                        if not result['df'].empty:
                            st.dataframe(
                                result['df'].head(10),
                                use_container_width=True,
                                hide_index=True
                            )
                            if len(result['df']) > 10:
                                st.caption(f"Viser 10 af {len(result['df'])} rækker")
                        else:
                            st.warning("Ingen data fundet i filen")

                    with col2:
                        # Reset BytesIO position
                        result['excel_data'].seek(0)
                        st.download_button(
                            label="⬇️ Download",
                            data=result['excel_data'],
                            file_name=result['output_name'],
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{result['filename']}"
                        )

            # Preview af samlet data
            st.divider()
            st.subheader("👁️ Samlet preview")

            all_dfs = [r['df'] for r in all_results if not r['df'].empty]
            if all_dfs:
                combined_df = pd.concat(all_dfs, ignore_index=True)
                st.dataframe(
                    combined_df,
                    use_container_width=True,
                    hide_index=True,
                    height=400
                )

                # Download samlet fil
                combined_excel = create_excel(combined_df, koerselsdato_str)
                st.download_button(
                    label="📥 Download samlet Excel-fil",
                    data=combined_excel,
                    file_name="rutelister_samlet.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    else:
        # Ingen filer - vis instruktioner
        st.markdown("""
        <div style="
            border: 2px dashed #7c3aed;
            border-radius: 12px;
            padding: 3rem;
            text-align: center;
            background: rgba(124, 58, 237, 0.1);
            margin: 2rem 0;
        ">
            <h3 style="color: #7c3aed;">📁 Drop RTF-filer her</h3>
            <p style="color: #94a3b8;">Eller klik på "Browse files" ovenfor</p>
        </div>
        """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
